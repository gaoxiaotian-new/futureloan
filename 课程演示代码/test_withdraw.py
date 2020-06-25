import os
import unittest
import ddt

from middleware import handler
from common import requests_handler
from common.excel_handler import ExcelHandler
from decimal import Decimal
from openpyxl.styles import PatternFill

# 初始化数据
env_data = handler.MiddleHandler()
logger = env_data.logger
cases = env_data.excel.read_data("withdraw")


@ddt.ddt
class TestWithdraw(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.token = env_data.token
        cls.member_id = env_data.member_id

    def setUp(self):
        self.db = env_data.db_class()

    def tearDown(self):
        self.db.close()

    @ddt.data(*cases)
    def test_withdraw(self, case_info):
        """测试提现接口"""
        data = case_info["data"]
        logger.info("正在执行第{}条用例：{}".format(case_info["case_id"], case_info["title"]))

        if "#member_id#" in data:
            data = data.replace("#member_id#", str(env_data.member_id))

        if "#amount#" in data:
            user_info = self.db.query("SELECT * FROM member WHERE id ={}".format(self.member_id))
            amount_leave = user_info["leave_amount"]
            # amount_new = "%.2f"%(amount_leave/2)
            amount_new = round(amount_leave / 10)
            data = data.replace("#amount#", str(amount_new))
            data = eval(data)
            data["amount"] = int(data["amount"])
            data = str(data)

        if "#amount_1#" in data:
            user_info = self.db.query("SELECT * FROM member WHERE id ={}".format(self.member_id))
            amount_leave = user_info["leave_amount"]
            amount_new = "%.1f" % (amount_leave / 10)
            data = data.replace("#amount_1#", amount_new)
            data = eval(data)
            data["amount"] = float(data["amount"])
            data = str(data)

        if "#amount_2#" in data:
            user_info = self.db.query("SELECT * FROM member WHERE id ={}".format(self.member_id))
            amount_leave = user_info["leave_amount"]
            amount_new = "%.2f" % (amount_leave / 10)
            data = data.replace("#amount_2#", amount_new)
            data = eval(data)
            data["amount"] = float(data["amount"])
            data = str(data)

        if case_info["check"]:
            env_data.recharge()

        headers = case_info["headers"]
        if "#token#" in headers:
            headers = headers.replace("#token#", self.token)

        # 访问接口前查询余额
        user_info = self.db.query("SELECT * FROM member WHERE id = {}".format(self.member_id))
        before_amount = user_info["leave_amount"]

        data = eval(data)
        resp = requests_handler.visit(
            url=env_data.yaml_data["host"] + case_info["url"],
            method=case_info["method"],
            json=data,
            headers=eval(headers)
        )

        data_path = os.path.join(env_data.config.DATA_PATH, "testcases.xlsx")
        try:
            expected = eval(case_info["expected"])
            for k, v in expected.items():
                self.assertTrue(v == resp[k])

            if expected["code"] == 0:
                user_info_after = self.db.query("SELECT * FROM member WHERE id = {}".format(self.member_id))
                after_amount = user_info_after["leave_amount"]
                self.assertTrue(
                    Decimal(str(before_amount)) - Decimal(str(data["amount"])) == Decimal(str(after_amount)))
            self.result = "PASS"
            logger.info("第{}条测试用例通过".format(case_info["case_id"]))
        except AssertionError as e:
            self.result = "FAIL"
            logger.error("第{}条测试用例无法通过：{}".format(case_info["case_id"], e))
            raise e
        finally:
            # if self.result == "FAIL":
            #     red_fill = PatternFill("solid", fgColor="FF0000")
            #     cell = ExcelHandler(data_path).get_sheet(sheet_name="withdraw").cell(row=case_info["case_id"] + 1,
            #                                                                          column=10)
            #     cell.fill = red_fill
            ExcelHandler(data_path).write(sheet_name="withdraw", row=case_info["case_id"] + 1, column=10,
                                          data=self.result)


if __name__ == '__main__':
    unittest.main()
